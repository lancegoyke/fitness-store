"""Parse a Google-Sheets-exported template workbook into a ``build_block`` spec.

Phase 3 of the spreadsheet-parity plan (§5 "Import — validate, don't
bulk-load"): a template ``.xlsx`` (one of Lance's ~57 single-sheet program
templates, exported via the Drive connector — fixtures under
``docs/meso/fixtures/templates/``) parses into the exact dict shape
``seed_meso_demo.build_block`` materializes, so the importer command is just
Plan + Mesocycle bookkeeping around this module.

The grid (see ``docs/meso/fixtures/sheet-structure.md`` and
``templates/design-grounding.md``):

- One **program tab** per workbook (some bundle metadata tabs — Athlete /
  Warm Up / FAQ / Periodization — and 102 carries a *hidden* legacy tab that
  must be skipped). The program tab is the visible sheet whose header row
  carries both an ``Exercise`` label and ``Week N`` columns (the Warm Up tab
  has an Exercise column but no weeks).
- **Header rows repeat per Day section** (``Day 1 | Exercise | Tempo |
  Week 1..N | Coach Comments | Athlete Comments | Rest``). Column letters
  drift between template generations — columns are resolved from the header
  LABELS per section, never hardcoded.
- An **exercise block** = the row under a header (name/tempo/coach-comment/
  rest MERGED down the block; the merge range gives the block extent) plus
  blank set-detail rows (the athlete's log area — skipped). A non-empty row
  inside the block extent (e.g. the newer templates' ``RPE 8 | RPE 9 | …``
  sub-row) folds into each week's cell as a sub-line (§2.3 / D3).
- **Non-grid rows**: the row-1 banner, ``Date:`` rows, and ``END OF WEEK``
  footers are skipped (recorded in the report); full-width separator rows
  *inside* a day (601's ``Rest 5 minutes``) import as a freeform exercise row
  with no cells — a row need not be an exercise (§2.2).

Everything maps VERBATIM — cell text, superset prefixes (``A)``/``C1)``),
packed circuit cells — structure is derived later by ``parsing.py``, never at
import. Defensive throughout: unknown structure is skipped and reported,
never raised mid-sheet.
"""

import re
from dataclasses import dataclass
from dataclasses import field

from openpyxl import load_workbook

# ``ExerciseSlot.name`` is a CharField(max_length=255); a packed circuit cell
# (601 Day 6) approaches it, so clip + report rather than crash the insert.
MAX_NAME_LENGTH = 255

_WEEK_RE = re.compile(r"^Week\s+(\d+)$", re.IGNORECASE)
_DAY_RE = re.compile(r"^Day\s+(\d+)$", re.IGNORECASE)


class SheetImportError(Exception):
    """The workbook has no parseable program tab."""


@dataclass
class SkippedRow:
    """One non-imported row, for the command's report."""

    row: int
    reason: str
    preview: str


@dataclass
class ParsedBlock:
    """One workbook's program grid, ready for ``build_block``.

    ``block_spec`` is the ``SAMPLE_PLAN``-mesocycle-shaped dict
    (``{"days": [...], "weeks": [...]}``); the rest is metadata for naming
    the Mesocycle and reporting the import.
    """

    tab: str
    week_count: int
    block_spec: dict
    skipped: list[SkippedRow] = field(default_factory=list)

    @property
    def day_count(self):
        return len(self.block_spec.get("days", []))

    @property
    def exercise_count(self):
        return sum(len(d.get("exercises", [])) for d in self.block_spec.get("days", []))

    @property
    def cell_count(self):
        """Non-empty imported cells: line-0 texts plus sub-line texts."""
        count = 0
        for week in self.block_spec.get("weeks", []):
            for cells in week.get("cells", {}).values():
                for cell in cells:
                    if cell.get("text"):
                        count += 1
                    count += sum(1 for line in cell.get("lines", []) if line)
        return count


def coerce_text(value):
    """A cell value as verbatim text; numerics lose the float artifact.

    Google Sheets stores the newer templates' numeric tempos as floats —
    openpyxl reads ``201`` as ``201.0`` — so integral floats coerce back to
    their integer spelling. Everything else is ``str()`` verbatim.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_columns(row_cells):
    """Resolve a header row's label → column map, or None if not a header.

    A program header row carries an ``Exercise`` label and at least one
    ``Week N`` column (the Warm Up tab's Exercise/Sets/Reps header has no
    weeks, so it never matches). Returns ``{"exercise": col, "tempo": col?,
    "note": col?, "rest": col?, "weeks": [(index, col), ...]}``.
    """
    columns = {"weeks": []}
    for cell in row_cells:
        if not isinstance(cell.value, str):
            continue
        label = cell.value.strip()
        if label == "Exercise":
            columns["exercise"] = cell.column
        elif label == "Tempo":
            columns["tempo"] = cell.column
        elif label == "Coach Comments":
            columns["note"] = cell.column
        elif label == "Rest":
            columns["rest"] = cell.column
        else:
            match = _WEEK_RE.match(label)
            if match:
                columns["weeks"].append((int(match.group(1)), cell.column))
    if "exercise" in columns and columns["weeks"]:
        columns["weeks"].sort()
        return columns
    return None


def _find_program_sheet(workbook):
    """The visible sheet containing a program grid, or None.

    Hidden sheets are never candidates (102's legacy ``Program`` tab); a
    metadata tab (Athlete / Warm Up / FAQ / Periodization) has no
    ``Exercise + Week N`` header row, so it never matches either.
    """
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            continue
        for row_cells in sheet.iter_rows():
            if _header_columns(row_cells) is not None:
                return sheet
    return None


def _merge_maps(sheet):
    """(top-left → merge range) and (any coordinate → merge range) maps."""
    by_top = {}
    by_coord = {}
    for merged in sheet.merged_cells.ranges:
        by_top[(merged.min_row, merged.min_col)] = merged
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                by_coord[(row, col)] = merged
    return by_top, by_coord


def parse_workbook(path):
    """Parse one template workbook into a :class:`ParsedBlock`.

    Raises :class:`SheetImportError` when no visible sheet carries a program
    grid; anything unrecognized *within* the grid is skipped + reported, never
    raised.
    """
    workbook = load_workbook(path, data_only=True)
    sheet = _find_program_sheet(workbook)
    if sheet is None:
        raise SheetImportError(
            f"{path}: no visible sheet with an 'Exercise' + 'Week N' header row."
        )
    return _parse_sheet(sheet)


def _parse_sheet(sheet):
    grid = list(sheet.iter_rows())
    merge_by_top, merge_by_coord = _merge_maps(sheet)

    # Day sections: every repeated header row, with its own label→column map
    # (column letters drift between template generations — resolve per section).
    sections = []
    for row_cells in grid:
        columns = _header_columns(row_cells)
        if columns is None:
            continue
        row = row_cells[0].row
        day_number = None
        for cell in row_cells:
            if isinstance(cell.value, str):
                match = _DAY_RE.match(cell.value.strip())
                if match:
                    day_number = int(match.group(1))
                    break
        sections.append({"row": row, "columns": columns, "day_number": day_number})

    skipped = []
    days = []
    # cells_by_day[day_number] = per-exercise list of {week_index: {"text", "lines"}}
    cells_by_day = {}
    week_count = max(
        (w for section in sections for w, _ in section["columns"]["weeks"]),
        default=0,
    )

    def report(row, reason, cells):
        preview = " | ".join(
            coerce_text(c.value).replace("\n", " ")[:40] for c in cells if c.value
        )
        skipped.append(SkippedRow(row=row, reason=reason, preview=preview[:120]))

    # Rows before the first header: the full-width banner + a bare ``Date:``
    # row on the oldest generation — outside any grid, so skip + report.
    first_header = sections[0]["row"]
    for row_cells in grid[: first_header - 1]:
        if any(c.value is not None for c in row_cells):
            is_date = any(
                isinstance(c.value, str) and c.value.strip() == "Date:"
                for c in row_cells
            )
            report(
                row_cells[0].row,
                "date row" if is_date else "banner / pre-grid row",
                row_cells,
            )

    for section_index, section in enumerate(sections):
        columns = section["columns"]
        day_number = section["day_number"]
        if day_number is None:
            # Header without a recognizable ``Day N`` label — number it after
            # the previous section rather than dropping the whole day.
            day_number = (days[-1]["day_number"] + 1) if days else section_index + 1
        end = (
            sections[section_index + 1]["row"] - 1
            if section_index + 1 < len(sections)
            else sheet.max_row
        )
        exercises, day_cells = _parse_section(
            grid=grid,
            start=section["row"] + 1,
            end=end,
            columns=columns,
            merge_by_top=merge_by_top,
            merge_by_coord=merge_by_coord,
            report=report,
        )
        days.append(
            {
                "day_number": day_number,
                "name": f"Day {day_number}",
                "exercises": exercises,
            }
        )
        cells_by_day[day_number] = day_cells

    weeks = []
    for index in range(1, week_count + 1):
        cells = {}
        for day in days:
            day_number = day["day_number"]
            cells[day_number] = [
                {
                    "text": per_week.get(index, {}).get("text", ""),
                    "lines": per_week.get(index, {}).get("lines", []),
                }
                for per_week in cells_by_day[day_number]
            ]
        weeks.append(
            {
                "index": index,
                # The importer lands the coach on week 1, like a fresh block.
                "is_current": index == 1,
                "cells": cells,
            }
        )

    return ParsedBlock(
        tab=sheet.title,
        week_count=week_count,
        block_spec={"days": days, "weeks": weeks},
        skipped=skipped,
    )


def _parse_section(*, grid, start, end, columns, merge_by_top, merge_by_coord, report):
    """One Day section's exercise rows + per-week cells.

    Returns ``(exercises, day_cells)`` — ``exercises`` in ``build_block``'s
    ``"exercises"`` shape, ``day_cells`` a parallel list of
    ``{week_index: {"text": str, "lines": [str, ...]}}`` per row.
    """
    exercise_col = columns["exercise"]
    week_cols = columns["weeks"]
    exercises = []
    day_cells = []
    row = start
    while row <= end:
        if row - 1 >= len(grid):
            break
        row_cells = grid[row - 1]
        values = {c.column: c.value for c in row_cells}
        name_value = values.get(exercise_col)

        if name_value is None:
            # An empty merged name block (the templates keep spare, pre-merged
            # exercise slots) consumes its whole extent silently; anything
            # else non-empty here is unrecognized structure.
            merged = merge_by_top.get((row, exercise_col))
            if merged is not None and merged.min_row == merged.max_row == row:
                merged = None  # single-row merge with no value — fall through
            if merged is not None:
                row = merged.max_row + 1
                continue
            if any(v is not None for v in values.values()):
                if any(
                    isinstance(v, str) and v.strip() == "Date:" for v in values.values()
                ):
                    report(row, "date row", row_cells)
                else:
                    report(row, "unrecognized row", row_cells)
            row += 1
            continue

        name = coerce_text(name_value)

        # Full-width merged rows: the ``END OF WEEK`` footer is chrome (skip);
        # any other (601's ``Rest 5 minutes``) is a real separator the coach
        # typed — import it as a freeform, cell-less row (§2.2: a row need not
        # be an exercise).
        merged = merge_by_coord.get((row, exercise_col))
        if (
            merged is not None
            and merged.min_row == merged.max_row == row
            and merged.max_col >= max(col for _, col in week_cols)
        ):
            if name.upper().startswith("END OF WEEK"):
                report(row, "end-of-week footer", row_cells)
            else:
                exercises.append({"name": name[:MAX_NAME_LENGTH]})
                day_cells.append({})
            row += 1
            continue

        # A real exercise block: its extent comes from the name cell's merge
        # range (name/tempo/note/rest merge down the block); an unmerged name
        # cell is a one-row block.
        block_top = merge_by_top.get((row, exercise_col))
        block_end = block_top.max_row if block_top is not None else row
        block_end = min(block_end, end)

        if len(name) > MAX_NAME_LENGTH:
            report(row, "name clipped to 255 chars", row_cells)
            name = name[:MAX_NAME_LENGTH]

        exercise = {"name": name}
        tempo = coerce_text(values.get(columns.get("tempo")))
        rest = coerce_text(values.get(columns.get("rest")))
        note = coerce_text(values.get(columns.get("note")))
        if tempo:
            exercise["tempo"] = tempo
        if rest:
            exercise["rest"] = rest
        if note:
            exercise["note"] = note

        per_week = {
            index: {"text": coerce_text(values.get(col)), "lines": []}
            for index, col in week_cols
        }

        # Sub-line rows (§2.3 / D3): any non-empty week-column row inside the
        # block extent — the newer templates' RPE row — folds into each week's
        # cell as a sub-line, aligned across weeks. Empty rows are the
        # athlete's set-detail log area (blank in a template) — skipped.
        for sub_row in range(row + 1, block_end + 1):
            if sub_row - 1 >= len(grid):
                break
            sub_values = {c.column: c.value for c in grid[sub_row - 1]}
            week_values = {
                index: coerce_text(sub_values.get(col)) for index, col in week_cols
            }
            if not any(week_values.values()):
                continue
            for index, text in week_values.items():
                per_week[index]["lines"].append(text)
        # Trailing blank sub-lines per week are alignment padding — drop them.
        for cell in per_week.values():
            while cell["lines"] and not cell["lines"][-1]:
                cell["lines"].pop()

        exercises.append(exercise)
        day_cells.append(per_week)
        row = block_end + 1

    return exercises, day_cells
