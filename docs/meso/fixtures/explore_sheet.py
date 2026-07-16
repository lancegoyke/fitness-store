#!/usr/bin/env python3
"""Reusable explorer for the Lance program workbook.

Opens the decoded xlsx with openpyxl (data_only=True so formulas resolve to
cached values), captures each tab's used grid (trailing empty rows/cols
trimmed), records merged-cell ranges, and dumps everything to sheet-dump.json.

Run ephemerally (do NOT add openpyxl to the project):
    uv run --with openpyxl python explore_sheet.py
"""

import json
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "lance-program.xlsx")
JSON_OUT = os.path.join(HERE, "sheet-dump.json")


def cell_to_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def trim_grid(grid):
    """Trim trailing all-empty rows and columns from a list-of-rows grid."""
    # trim trailing empty rows
    while grid and all(c == "" for c in grid[-1]):
        grid.pop()
    if not grid:
        return grid
    # trim trailing empty columns
    max_used_col = 0
    for row in grid:
        for idx in range(len(row) - 1, -1, -1):
            if row[idx] != "":
                if idx + 1 > max_used_col:
                    max_used_col = idx + 1
                break
    return [row[:max_used_col] for row in grid]


def main():
    wb = load_workbook(XLSX, data_only=True, read_only=False)
    out = {
        "workbook": os.path.basename(XLSX),
        "sheet_count": len(wb.sheetnames),
        "sheet_order": list(wb.sheetnames),
        "sheets": [],
    }

    for name in wb.sheetnames:
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        grid = []
        for row in ws.iter_rows(
            min_row=1, max_row=max_row, max_col=max_col, values_only=True
        ):
            grid.append([cell_to_str(v) for v in row])
        grid = trim_grid(grid)

        non_empty_rows = sum(1 for r in grid if any(c != "" for c in r))

        merged = [str(rng) for rng in ws.merged_cells.ranges]

        trimmed_rows = len(grid)
        trimmed_cols = max((len(r) for r in grid), default=0)

        out["sheets"].append(
            {
                "name": name,
                "max_row": max_row,
                "max_col": max_col,
                "max_col_letter": get_column_letter(max_col) if max_col else "",
                "trimmed_rows": trimmed_rows,
                "trimmed_cols": trimmed_cols,
                "non_empty_rows": non_empty_rows,
                "merged_ranges": merged,
                "grid": grid,
            }
        )

    with open(JSON_OUT, "w") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)

    # concise console summary
    print(f"workbook: {out['workbook']}  tabs: {out['sheet_count']}")
    print("-" * 78)
    print(f"{'#':>2}  {'tab name':32}  {'rows':>5} {'cols':>4} {'ne_rows':>7}  merges")
    for i, s in enumerate(out["sheets"], 1):
        print(
            f"{i:>2}  {s['name'][:32]:32}  {s['trimmed_rows']:>5} "
            f"{s['trimmed_cols']:>4} {s['non_empty_rows']:>7}  "
            f"{len(s['merged_ranges'])}"
        )
    print("-" * 78)
    print(f"wrote {JSON_OUT} ({os.path.getsize(JSON_OUT)} bytes)")


if __name__ == "__main__":
    main()
