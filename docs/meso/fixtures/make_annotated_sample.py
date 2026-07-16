"""Derive an anonymized annotated-program fixture from a real client workbook.

The client's identity (name, contact) lives only in the workbook's metadata tabs,
which are NOT copied here — we extract only the single program tab `415 - 0626`
(pure workout data: exercises, prescriptions, executions, RPE, coach cues) and drop
the training-date values, leaving no PII. Kept for provenance/reproducibility.

Run: uv run --with openpyxl python make_annotated_sample.py <raw_client.xlsx> <out.xlsx>
"""

import datetime
import sys

import openpyxl


def find_program_tab(wb):
    for ws in wb.worksheets:
        b68 = str(ws["B68"].value or "")
        if "Turkish get up" in b68 and ws["D124"].value == "DB pullover":
            return ws
    raise SystemExit("target program tab not found")


def main(src_path, out_path):
    src = openpyxl.load_workbook(src_path, data_only=True)
    ws = find_program_tab(src)

    out = openpyxl.Workbook()
    o = out.active
    o.title = "Annotated program sample"

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, (datetime.datetime, datetime.date)):
                continue  # drop training dates — anonymize
            o[cell.coordinate] = v

    for mr in ws.merged_cells.ranges:
        o.merge_cells(str(mr))

    out.save(out_path)
    print(
        f"wrote {out_path}: {o.max_row} rows x {o.max_column} cols, "
        f"{len(ws.merged_cells.ranges)} merges (dates dropped)"
    )


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
